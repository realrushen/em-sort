# coding=utf-8
from __future__ import annotations

from pathlib import Path
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from entities import Device, Schematic
from exceptions import UnsupportedTypeException, SheetDoesNotExistsException
from parser import Parser


class Sorter:
    # FIXME: make loading these constants from settings
    INPUT_DATA_COLUMN = 'A'

    def __init__(self, workbook: Optional[Workbook] = None):
        self._input_wb = workbook
        self._output_wb = Workbook()

        self.schematic = Schematic()
        self.parser = Parser(workbook, self.schematic.content)
        self._sheets_for_sort: List[str] = []

        # remove created by default sheet
        self._output_wb.remove(self._output_wb.active)

    def add_sheets(self, sheets_names: List[str]) -> None:
        for name in sheets_names:
            try:
                self.wb[name]
            except KeyError:
                raise SheetDoesNotExistsException(f'Worksheet {name} does not exist.')
            if name not in self._sheets_for_sort:
                self._sheets_for_sort.append(name)

    @property
    def wb(self) -> Optional[Workbook]:
        return self._input_wb

    @wb.setter
    def wb(self, workbook: Workbook) -> None:
        if isinstance(workbook, Workbook):
            self._input_wb = workbook
            self.parser.workbook = workbook
        else:
            raise UnsupportedTypeException

    def sort(self):
        self.parser.parse()
        for wire_section, devices in self.schematic.content.items():
            if wire_section in self._sheets_for_sort:
                for device in devices:
                    device.sort()

    def dump_circuitry(self) -> None:
        for wire_section, devices in self.schematic.content.items():
            worksheet = self._output_wb.create_sheet(wire_section)
            self._write_markers(worksheet=worksheet, devices=devices)

    def save_to_file(self, target_file_path: Path, in_place=False) -> None:
        if not in_place:
            target_file_path.with_name(f'{target_file_path.stem}_sorted')
        self._output_wb.save(target_file_path)

    @classmethod
    def reset(cls) -> Sorter:
        """Resets object to initial state"""
        return cls()

    @staticmethod
    def _write_markers(worksheet: Worksheet, devices: List[Device], column: int = 1) -> None:
        row = 1
        for device in devices:
            device_sell = worksheet.cell(row=row, column=column, value=device.name)
            device_sell.fill = PatternFill(fill_type='solid', start_color='00C0C0C0', end_color='00C0C0C0')
            row += 1
            for marker in device.markers:
                worksheet.cell(row=row, column=column, value=marker)
                row += 1
