from pathlib import Path

import openpyxl
import pytest

from sorter import Sorter

TEST_DATA_FOLDER = Path().cwd() / 'testdata'


@pytest.fixture
def clean_workbook():
    return openpyxl.Workbook()


@pytest.fixture
def example_schematic_workbook():
    return openpyxl.load_workbook(TEST_DATA_FOLDER / 'schematic1.xlsx')


@pytest.fixture
def sorted_example_schematic_workbook():
    return openpyxl.load_workbook(TEST_DATA_FOLDER / 'schematic1_sorted.xlsx')


@pytest.fixture
def expected_sorted_schematic(wire_sections_for_sort, sorted_example_schematic_workbook):
    sorter = Sorter(workbook=sorted_example_schematic_workbook)
    sorter.add_sheets(wire_sections_for_sort)

    sorter.sort()
    return sorter.schematic.content


@pytest.fixture
def wire_sections_for_sort():
    return ['1,0', '1,5', '2,5']
